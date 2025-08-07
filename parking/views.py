# Python standard library
from datetime import datetime, timedelta

# Django core
from django.contrib import messages
from django.contrib.auth import logout, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.views import LoginView
from django.db import IntegrityError, models
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import timezone
from django.utils.timezone import now

# Django database
from django.db.models import Avg, Count, F, Q, Sum
from django.db.models.functions import Extract, TruncDate

# Django views
from django.views.generic import CreateView, ListView, TemplateView, UpdateView
from django.views.generic.edit import DeleteView
from django.utils.decorators import method_decorator
from django.views import View

# Local imports
from .forms import CategoryForm, ParkingLotForm, ParkingTicketForm, TenantForm, TenantCreateForm, TenantEditForm, TenantUserForm, TenantUserEditForm, TenantUserPasswordForm, TenantLoginForm, CajaTurnoAperturaForm, CajaMovimientoForm, CajaTurnoCierreForm
from .models import ParkingLot, ParkingTicket, VehicleCategory, Tenant, UserProfile, CajaTurno, CajaMovimiento
from .decorators import tenant_required, permission_required, admin_required, operator_required, cajero_required

import io
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib import colors


class CustomLoginView(LoginView):
    """
    Vista personalizada de login que usa TenantLoginForm para usuarios normales
    y AuthenticationForm para superadmin
    """
    template_name = 'registration/login.html'
    
    def get_form_class(self):
        # Si hay un parámetro next que apunte a superadmin, usar formulario simple
        next_url = self.request.GET.get('next', '')
        if 'superadmin' in next_url or self.request.GET.get('superadmin') == 'true':
            from django.contrib.auth.forms import AuthenticationForm
            return AuthenticationForm
        return TenantLoginForm
    
    def form_valid(self, form):
        next_url = self.request.GET.get('next', '')
        
        # Si es superadmin, usar autenticación normal
        if 'superadmin' in next_url or self.request.GET.get('superadmin') == 'true':
            return super().form_valid(form)
        
        # Para usuarios normales, usar lógica multitenant
        if hasattr(form, 'cleaned_data') and 'tenant' in form.cleaned_data:
            tenant = form.cleaned_data.get('tenant')
            user = form.cleaned_data.get('user')
            
            if tenant and user:
                # Autenticar al usuario
                login(self.request, user)
                
                # Guardar el tenant en la sesión
                self.request.session['tenant_id'] = tenant.id
                self.request.session['tenant_name'] = tenant.name
                
                # Redirigir al dashboard
                return redirect('dashboard')
        
        return super().form_valid(form)
    
    def form_invalid(self, form):
        next_url = self.request.GET.get('next', '')
        
        # Si es superadmin, usar mensaje de error normal
        if 'superadmin' in next_url or self.request.GET.get('superadmin') == 'true':
            messages.error(self.request, "Credenciales inválidas.")
        else:
            messages.error(self.request, "Error en las credenciales. Verifique su correo electrónico y contraseña.")
        
        return super().form_invalid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        next_url = self.request.GET.get('next', '')
        
        # Determinar si es login de superadmin
        is_superadmin_login = 'superadmin' in next_url or self.request.GET.get('superadmin') == 'true'
        context['is_superadmin_login'] = is_superadmin_login
        
        return context


def custom_logout(request):
    """Vista personalizada de logout"""
    logout(request)
    # Limpiar datos de tenant de la sesión
    if 'tenant_id' in request.session:
        del request.session['tenant_id']
    if 'tenant_name' in request.session:
        del request.session['tenant_name']
    return redirect('login')


def pagina_inicial(request):
    """Página inicial que redirige según el tipo de usuario"""
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect('superadmin-dashboard')
        else:
            return redirect('dashboard')
    else:
        return redirect('login')


class ParkingLotUpdateView(UpdateView):
    model = ParkingLot
    template_name = 'parking/parking_lot_form.html'
    fields = ['empresa', 'nit', 'telefono', 'direccion']
    success_url = reverse_lazy('dashboard')


@method_decorator(login_required, name='dispatch')
@method_decorator(tenant_required, name='dispatch')
@method_decorator(admin_required, name='dispatch')
class CategoryListView(ListView):
    model = VehicleCategory
    template_name = 'parking/category_list_simple.html'

    def get_queryset(self):
        return VehicleCategory.objects.filter(tenant=self.request.tenant)


@method_decorator(login_required, name='dispatch')
@method_decorator(tenant_required, name='dispatch')
@method_decorator(admin_required, name='dispatch')
class CategoryCreateView(CreateView):
    model = VehicleCategory
    form_class = CategoryForm
    template_name = 'parking/category_form.html'
    success_url = reverse_lazy('category-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['tenant'] = self.request.tenant
        return kwargs

    def form_valid(self, form):
        form.instance.tenant = self.request.tenant
        messages.success(self.request, "Categoría creada con éxito.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Hubo un error al crear la categoría.")
        return super().form_invalid(form)


@method_decorator(login_required, name='dispatch')
@method_decorator(tenant_required, name='dispatch')
@method_decorator(admin_required, name='dispatch')
class CategoryEditView(UpdateView):
    model = VehicleCategory
    form_class = CategoryForm
    template_name = 'parking/category_form.html'
    success_url = reverse_lazy('category-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['tenant'] = self.request.tenant
        return kwargs

    def get_queryset(self):
        return VehicleCategory.objects.filter(tenant=self.request.tenant)

    def form_valid(self, form):
        form.instance.tenant = self.request.tenant
        messages.success(self.request, "Categoría actualizada con éxito.")
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, "Hubo un error al actualizar la categoría.")
        return super().form_invalid(form)


@method_decorator(login_required, name='dispatch')
@method_decorator(tenant_required, name='dispatch')
@method_decorator(operator_required, name='dispatch')
class VehicleEntryView(CreateView):
    model = ParkingTicket
    form_class = ParkingTicketForm
    template_name = 'parking/vehicle_entry.html'
    success_url = reverse_lazy('print-ticket')

    def form_valid(self, form):
        try:
            form.instance.tenant = self.request.tenant
            response = super().form_valid(form)
            self.request.session['ticket_id'] = str(self.object.id)
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                from django.urls import reverse
                return JsonResponse({
                    'success': True,
                    'redirect_url': reverse('print-ticket')
                })
            return response
        except IntegrityError:
            messages.error(self.request, 'Este vehículo ya se encuentra en el estacionamiento.')
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'error': 'Este vehículo ya se encuentra en el estacionamiento.'}, status=400)
            return self.form_invalid(form)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Pasar el tenant al formulario para que configure los campos correctamente
        if hasattr(form, 'fields') and 'category' in form.fields:
            form.fields['category'].queryset = VehicleCategory.objects.filter(tenant=self.request.tenant)
        return form


@login_required
@tenant_required
@cajero_required
def vehicle_exit(request):
    print(f"vehicle_exit called - Method: {request.method}")
    
    if request.method == 'POST':
        identifier = request.POST.get('identifier')
        print(f"Searching for vehicle with identifier: {identifier}")
        
        ticket = ParkingTicket.objects.filter(
            tenant=request.tenant,
            placa__iexact=identifier,
            exit_time=None
        ).first()
        
        print(f"Ticket found: {ticket is not None}")

        if ticket:
            try:
                # Solo retornar datos, NO registrar salida ni movimiento
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    print("Processing AJAX request")
                    
                    duration = ticket.get_current_duration() if not ticket.exit_time else ticket.get_duration()
                    if isinstance(duration, dict):
                        duration_str = f"{duration['hours']}h {duration['minutes']}m"
                    else:
                        duration_str = f"{duration}h"
                    
                    # Calcular el monto total
                    total_amount = ticket.calculate_current_fee() if not ticket.exit_time else ticket.calculate_fee()
                    print(f"Total amount calculated: {total_amount}")
                    
                    # Obtener desglose de impuestos
                    breakdown = ticket.get_formatted_breakdown()
                    print(f"Breakdown: {breakdown}")
                    
                    response_data = {
                        'amount': total_amount,
                        'duration': duration_str,
                        'placa': ticket.placa,
                        'entry_time': ticket.entry_time.strftime('%Y-%m-%d %H:%M:%S'),
                        'ticket_id': str(ticket.id),
                        'tax_breakdown': breakdown
                    }
                    print(f"Returning response: {response_data}")
                    
                    return JsonResponse(response_data)
            except Exception as e:
                print(f"Error in vehicle_exit: {str(e)}")
                import traceback
                traceback.print_exc()
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'error': str(e)}, status=500)
                messages.error(request, f'Error: {str(e)}')
                return redirect('vehicle-exit')

        # Si no se encuentra el ticket
        print("Vehicle not found")
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'error': 'Vehículo no encontrado'}, status=404)
        messages.error(request, 'Vehículo no encontrado')
        return redirect('vehicle-exit')

    # Para solicitudes GET
    placa = request.GET.get('placa', '')
    print(f"GET request - placa: {placa}")
    return render(request, 'parking/vehicle_exit.html', {
        'placa': placa,
    })


@login_required
@tenant_required
@cajero_required
def print_exit_ticket(request):
    if request.method == 'POST':
        print(f"print_exit_ticket POST recibido")
        print(f"POST data: {request.POST}")
        print(f"Headers: {dict(request.headers)}")
        
        ticket_id = request.POST.get('ticket_id')
        amount_received = request.POST.get('amount_received')
        forma_pago = request.POST.get('forma_pago', 'efectivo')
        
        print(f"ticket_id: {ticket_id}")
        print(f"amount_received: {amount_received}")
        print(f"forma_pago: {forma_pago}")

        if ticket_id and amount_received:
            try:
                ticket = ParkingTicket.objects.get(id=ticket_id, tenant=request.tenant)
                amount_received = float(amount_received)
                amount_paid = float(ticket.calculate_fee())
                change = amount_received - amount_paid

                # Registrar la salida y el movimiento de caja aquí
                if not ticket.exit_time:
                    ticket.exit_time = timezone.now()
                    ticket.amount_paid = amount_paid
                    ticket.amount_received = amount_received  # Guardar recibido
                    ticket.change = change  # Guardar vuelto
                    ticket.forma_pago = forma_pago  # Guardar forma de pago
                    ticket.save()

                    # Registrar movimiento de caja
                    turno = CajaTurno.objects.filter(tenant=request.tenant, estado='abierto').first()
                    if turno and ticket.amount_paid:
                        CajaMovimiento.objects.create(
                            caja_turno=turno,
                            tipo='Ingreso',
                            categoria='parqueo',
                            monto=ticket.amount_paid,
                            descripcion=f'Ingreso por parqueo - Placa: {ticket.placa}',
                            forma_pago=forma_pago,
                            usuario=request.user,
                            tipo_vehiculo=ticket.category.name if ticket.category else ''
                        )

                parking_lot = ParkingLot.objects.filter(tenant=request.tenant).first()

                # Si es AJAX, devolver JSON con la URL de impresión
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    from django.urls import reverse
                    return JsonResponse({
                        'success': True,
                        'redirect_url': reverse('print-exit-ticket') + f'?ticket_id={ticket.id}'
                    })

                return render(request, 'parking/print_exit_ticket.html', {
                    'ticket': ticket,
                    'parking_lot': parking_lot,
                    'amount_received': amount_received,
                    'change': change,
                    'current_time': timezone.now(),
                })
            except ParkingTicket.DoesNotExist:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': 'Ticket no encontrado'}, status=404)
                messages.error(request, 'Ticket no encontrado')
                return redirect('dashboard')
            except ValueError:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': 'El monto recibido no es válido'}, status=400)
                messages.error(request, 'El monto recibido no es válido')
                return redirect('vehicle-exit')
            except Exception as e:
                import traceback
                print(f"Error en print_exit_ticket: {str(e)}")
                print(traceback.format_exc())
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': f'Error al imprimir ticket: {str(e)}'}, status=500)
                messages.error(request, f'Error al imprimir ticket: {str(e)}')
                return redirect('dashboard')

    # NUEVO: Permitir GET para mostrar el ticket de salida
    if request.method == 'GET':
        ticket_id = request.GET.get('ticket_id')
        if ticket_id:
            try:
                ticket = ParkingTicket.objects.get(id=ticket_id, tenant=request.tenant)
                parking_lot = ParkingLot.objects.filter(tenant=request.tenant).first()
                return render(request, 'parking/print_exit_ticket.html', {
                    'ticket': ticket,
                    'parking_lot': parking_lot,
                    'amount_received': ticket.amount_received or ticket.amount_paid or 0,
                    'change': ticket.change or 0,
                    'current_time': timezone.now(),
                })
            except ParkingTicket.DoesNotExist:
                messages.error(request, 'Ticket no encontrado')
                return redirect('dashboard')
        messages.warning(request, 'No se especificó un ticket para imprimir')
        return redirect('dashboard')

    messages.warning(request, 'No se especificó un ticket para imprimir')
    return redirect('dashboard')


@login_required
@tenant_required
@operator_required
def dashboard(request):
    """Dashboard principal del sistema"""
    if not request.user.is_authenticated:
        return redirect('login')
    
    # Si es superusuario, redirigir al dashboard de superadmin
    if request.user.is_superuser:
        return redirect('superadmin-dashboard')
    
    # Obtener estadísticas del tenant actual
    tenant = request.tenant
    
    # Estadísticas de tickets
    total_tickets = ParkingTicket.objects.filter(tenant=tenant).count()
    active_tickets = ParkingTicket.objects.filter(tenant=tenant, exit_time__isnull=True).count()
    completed_tickets = ParkingTicket.objects.filter(tenant=tenant, exit_time__isnull=False).count()
    
    # Ingresos del día
    today = timezone.localtime(timezone.now()).date()
    today_income = ParkingTicket.objects.filter(
        tenant=tenant,
        exit_time__date=today,
        amount_paid__isnull=False
    ).aggregate(total=Sum('amount_paid'))['total'] or 0
    
    # Si no hay ingresos de hoy, intentar con todos los tickets completados
    if today_income == 0:
        today_income = ParkingTicket.objects.filter(
            tenant=tenant,
            exit_time__isnull=False,
            amount_paid__isnull=False
        ).aggregate(total=Sum('amount_paid'))['total'] or 0
    
    print(f"DEBUG Dashboard - Today: {today}")
    print(f"DEBUG Dashboard - Today income: {today_income}")
    print(f"DEBUG Dashboard - Active tickets: {active_tickets}")
    print(f"DEBUG Dashboard - Total tickets: {total_tickets}")
    
    # Tickets del día
    today_tickets = ParkingTicket.objects.filter(
        tenant=tenant,
        entry_time__date=today
    ).count()
    
    # Categorías más usadas
    popular_categories = VehicleCategory.objects.filter(tenant=tenant).annotate(
        ticket_count=Count('parkingticket')
    ).order_by('-ticket_count')[:5]
    
    # Vehículos activos actualmente
    active_vehicles = ParkingTicket.objects.filter(
        tenant=tenant,
        exit_time__isnull=True
    ).select_related('category').order_by('-entry_time')
    
    # Estadísticas por categoría
    category_stats = VehicleCategory.objects.filter(tenant=tenant).annotate(
        total_tickets=Count('parkingticket'),
        active_tickets=Count('parkingticket', filter=Q(parkingticket__exit_time__isnull=True)),
        completed_tickets=Count('parkingticket', filter=Q(parkingticket__exit_time__isnull=False))
    ).order_by('-total_tickets')
    
    print(f"DEBUG Dashboard - Category stats count: {category_stats.count()}")
    for cat in category_stats:
        print(f"DEBUG Dashboard - Category {cat.name}: {cat.total_tickets} tickets")
    
    # Estadísticas de los últimos 7 días
    seven_days_ago = timezone.now() - timedelta(days=7)
    weekly_stats = ParkingTicket.objects.filter(
        tenant=tenant,
        entry_time__gte=seven_days_ago
    ).aggregate(
        total_entries=Count('id'),
        total_completed=Count('id', filter=Q(exit_time__isnull=False)),
        total_income=Sum('amount_paid')
    )
    
    # Si no hay estadísticas semanales, usar todos los tickets
    if not weekly_stats['total_entries']:
        weekly_stats = ParkingTicket.objects.filter(tenant=tenant).aggregate(
            total_entries=Count('id'),
            total_completed=Count('id', filter=Q(exit_time__isnull=False)),
            total_income=Sum('amount_paid')
        )
    
    print(f"DEBUG Dashboard - Weekly stats: {weekly_stats}")
    print(f"DEBUG Dashboard - Seven days ago: {seven_days_ago}")
    
    # Calcular ingresos por categoría de forma más simple
    category_list = []
    for cat in category_stats:
        revenue = ParkingTicket.objects.filter(
            tenant=tenant,
            category=cat,
            amount_paid__isnull=False
        ).aggregate(total=Sum('amount_paid'))['total'] or 0
        
        category_list.append({
            'category__name': cat.name,
            'count': cat.total_tickets,
            'revenue': revenue
        })
        print(f"DEBUG Dashboard - Category {cat.name}: {cat.total_tickets} tickets, ${revenue} revenue")

    # Preparar estadísticas para el template
    stats = {
        'daily': {
            'total_vehicles': weekly_stats['total_entries'] or 0,
            'total_income': today_income,  # Ingresos de hoy
            'total_revenue': weekly_stats['total_income'] or 0,
        },
        'active_vehicles': active_tickets,
        'active_vehicles_list': active_vehicles,
        'category': category_list
    }
    
    print(f"DEBUG Dashboard - Final stats: {stats}")
    
    # Estadísticas diarias (últimos 7 días)
    daily_stats = []
    for i in range(7):
        date = timezone.localtime(timezone.now()).date() - timedelta(days=i)
        day_income = ParkingTicket.objects.filter(
            tenant=tenant,
            exit_time__date=date,
            amount_paid__isnull=False
        ).aggregate(total=Sum('amount_paid'))['total'] or 0
        
        daily_stats.append({
            'date': date.strftime('%d/%m/%Y'),
            'revenue': day_income
        })
    
    # Validar si hay turno de caja abierto
    turno_abierto = CajaTurno.objects.filter(tenant=tenant, estado='abierto').first()
    show_apertura_alert = False
    if not turno_abierto:
        show_apertura_alert = True

    context = {
        'stats': stats,
        'daily_stats': daily_stats,
        'current_time': timezone.now(),
        'show_apertura_alert': show_apertura_alert,
        'active_vehicles_count': active_tickets,  # Para la tarjeta de vehículos activos
        'available_spaces': 'N/A',  # Espacios disponibles (puedes configurar esto según tu lógica)
        'active_vehicles': active_vehicles,  # Lista de vehículos activos para la tabla
        'category_stats': category_stats,  # Estadísticas por categoría
    }
    
    return render(request, 'parking/dashboard.html', context)
@login_required
@tenant_required
def print_ticket(request):
    ticket_id = request.GET.get('ticket_id') or request.session.get('ticket_id')
    
    if ticket_id:
        try:
            ticket = ParkingTicket.objects.get(id=ticket_id, tenant=request.tenant)
            parking_lot = ParkingLot.objects.filter(tenant=request.tenant).first()
            
            if 'ticket_id' in request.session:
                del request.session['ticket_id']
            
            return render(request, 'parking/print_ticket.html', {
                'ticket': ticket,
                'parking_lot': parking_lot,
                'tenant': request.tenant,
                'is_reprint': bool(request.GET.get('ticket_id')),
                'current_time': timezone.now(),
                'duration': ticket.get_duration() if ticket.exit_time else ticket.get_current_duration(),
                'current_fee': ticket.amount_paid if ticket.exit_time else ticket.calculate_current_fee()
            })
        except ParkingTicket.DoesNotExist:
            messages.error(request, 'Ticket no encontrado')
            return redirect('dashboard')
        except Exception as e:
            messages.error(request, f'Error al imprimir ticket: {str(e)}')
            return redirect('dashboard')
    
    messages.warning(request, 'No se especificó un ticket para imprimir')
    return redirect('vehicle-entry')


@method_decorator(login_required, name='dispatch')
@method_decorator(tenant_required, name='dispatch')
@method_decorator(operator_required, name='dispatch')
class ReportView(TemplateView):
    template_name = 'parking/reports.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request = self.request
        # Filtros rápidos
        filtro_rapido = request.GET.get('quick_filter')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        # Obtener la fecha actual en la zona horaria configurada
        today = timezone.localtime(now()).date()
        
        if filtro_rapido == 'hoy':
            start_date = end_date = today
        elif filtro_rapido == 'semana':
            start_date = today - timedelta(days=today.weekday())
            end_date = today
        elif filtro_rapido == 'mes':
            start_date = today.replace(day=1)
            end_date = today
        elif filtro_rapido == 'anio':
            start_date = today.replace(month=1, day=1)
            end_date = today
        # Manejo de fechas
        if not start_date or not end_date:
            start_date = today
            end_date = today
        if isinstance(start_date, str):
            start_date = datetime.strptime(str(start_date), '%Y-%m-%d').date()
        if isinstance(end_date, str):
            end_date = datetime.strptime(str(end_date), '%Y-%m-%d').date()
        start_datetime = datetime.combine(start_date, datetime.min.time())
        end_datetime = datetime.combine(end_date, datetime.max.time())
        # Obtener tickets completados del tenant actual
        tickets = ParkingTicket.objects.filter(
            tenant=request.tenant,
            exit_time__isnull=False,
            exit_time__range=(start_datetime, end_datetime)
        ).exclude(amount_paid__isnull=True)
        # --- DESGLOSE DE INGRESOS POR FORMA DE PAGO (AGRUPAR POR TICKET) ---
        from parking.models import CajaMovimiento
        formas = [c[0] for c in CajaMovimiento.FORMA_PAGO_CHOICES]
        nombres = dict(CajaMovimiento.FORMA_PAGO_CHOICES)
        colores = {
            'efectivo': 'bg-green-600',
            'datafono': 'bg-blue-600',
            'transferencia': 'bg-indigo-600',
            'otro': 'bg-gray-700',
        }
        iconos = {
            'efectivo': 'payments',
            'datafono': 'credit_card',
            'transferencia': 'account_balance',
            'otro': 'attach_money',
        }
        ingresos_por_forma = {f: 0 for f in formas}
        total_general = 0
        for ticket in tickets:
            forma = ticket.forma_pago or 'efectivo'
            if forma not in ingresos_por_forma:
                forma = 'efectivo'
            ingresos_por_forma[forma] += float(ticket.amount_paid)
            total_general += float(ticket.amount_paid)
        # Lista profesional para el template
        formas_pago_list = []
        for forma in formas:
            valor = ingresos_por_forma[forma]
            porcentaje = (valor / total_general * 100) if total_general > 0 else 0
            formas_pago_list.append({
                'clave': forma,
                'nombre': nombres.get(forma, forma.title()),
                'valor': valor,
                'color': colores.get(forma, 'bg-gray-400'),
                'icono': iconos.get(forma, 'payments'),
                'porcentaje': porcentaje,
            })
        # --- RESUMEN GENERAL USANDO SOLO INGRESOS DE TICKETS ---
        summary = {
            'total_vehicles': tickets.count(),
            'total_revenue': total_general,  # Usar el total de tickets realmente cobrados
            'avg_duration': None,
            'avg_revenue': None,
        }
        # Calcular promedios solo si hay tickets
        if tickets.exists():
            summary['avg_duration'] = tickets.aggregate(avg=Avg(F('exit_time') - F('entry_time')))['avg']
            if summary['avg_duration'] is not None:
                summary['avg_duration'] = summary['avg_duration'].total_seconds() / 3600
            summary['avg_revenue'] = total_general / tickets.count() if tickets.count() > 0 else 0
        # --- ADVERTENCIA SI HAY DESCUBRE (NO DEBERÍA HABER) ---
        total_tickets = tickets.aggregate(total=Sum('amount_paid'))['total'] or 0
        # Asegurar que ambos sean Decimal para la resta
        from decimal import Decimal
        if not isinstance(total_tickets, Decimal):
            total_tickets = Decimal(str(total_tickets))
        if not isinstance(total_general, Decimal):
            total_general_decimal = Decimal(str(total_general))
        else:
            total_general_decimal = total_general
        descuadre = abs(total_tickets - total_general_decimal)
        advertencia_descuadre = None
        if descuadre > Decimal('1'):
            advertencia_descuadre = f"¡Atención! El total de ingresos por tickets (${total_tickets:,.2f}) no cuadra con el total de caja (${total_general:,.2f}). Revise movimientos manuales o errores de caja."
        # Estadísticas por categoría
        category_stats = []
        for stat in tickets.values('category__name').annotate(
            count=Count('id'),
            revenue=Sum('amount_paid'),
        ).order_by('-count'):
            category_tickets = tickets.filter(category__name=stat['category__name'])
            durations = [
                (ticket.exit_time - ticket.entry_time).total_seconds() / 3600
                for ticket in category_tickets
            ]
            avg_duration = sum(durations) / len(durations) if durations else 0
            stat['avg_duration'] = avg_duration
            category_stats.append(stat)
        # Estadísticas diarias
        daily_stats = list(tickets.annotate(
            date=TruncDate('exit_time')
        ).values('date').annotate(
            count=Count('id'),
            revenue=Sum('amount_paid')
        ).order_by('date'))
        # Vehículos más frecuentes
        frequent_vehicles = tickets.values('placa').annotate(
            visits=Count('id'),
            total_spent=Sum('amount_paid')
        ).order_by('-visits')[:10]
        context.update({
            'start_date': start_date,
            'end_date': end_date,
            'summary': summary,
            'category_stats': category_stats,
            'daily_stats': daily_stats,
            'frequent_vehicles': frequent_vehicles,
            'parking_lot': ParkingLot.objects.filter(tenant=request.tenant).first(),
            'formas_pago_list': formas_pago_list,
            'total_general_formas_pago': total_general,
            'quick_filter': filtro_rapido,
            'advertencia_descuadre': advertencia_descuadre,
        })
        return context

    def render_to_response(self, context, **response_kwargs):
        request = self.request
        export = request.GET.get('export')
        if export == 'excel':
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font
            from django.http import HttpResponse
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Reporte Ingresos'
            ws.append(['Fecha', 'Placa', 'Categoría', 'Monto', 'Forma de Pago'])
            for ticket in ParkingTicket.objects.filter(
                tenant=request.tenant,
                exit_time__isnull=False,
                exit_time__range=(datetime.combine(context['start_date'], datetime.min.time()), datetime.combine(context['end_date'], datetime.max.time()))
            ).exclude(amount_paid__isnull=True):
                ws.append([
                    ticket.exit_time.strftime('%Y-%m-%d %H:%M'),
                    ticket.placa,
                    ticket.category.name if ticket.category else '',
                    ticket.amount_paid,
                    dict(CajaMovimiento.FORMA_PAGO_CHOICES).get(ticket.forma_pago, ticket.forma_pago or ''),
                ])
            for col in range(1, 6):
                ws.column_dimensions[get_column_letter(col)].width = 18
            for cell in ws[1]:
                cell.font = Font(bold=True)
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=ReporteIngresos.xlsx'
            wb.save(response)
            return response
        elif export == 'pdf':
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
            from django.http import HttpResponse
            import io
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
            elements = []
            styles = getSampleStyleSheet()
            # Título premium
            title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=22, textColor=colors.HexColor('#2563eb'), alignment=TA_CENTER, spaceAfter=12)
            subtitle_style = ParagraphStyle('subtitle', parent=styles['Normal'], fontSize=12, textColor=colors.HexColor('#64748b'), alignment=TA_CENTER, spaceAfter=8)
            info_style = ParagraphStyle('info', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#334155'), alignment=TA_LEFT, spaceAfter=8)
            # Logo (opcional, si tienes uno en static)
            # elements.append(Image('static/logo.png', width=60, height=60))
            elements.append(Paragraph("Reporte de Ingresos - Parqueadero", title_style))
            elements.append(Paragraph(f"Fecha de generación: <b>{timezone.localtime(now()).strftime('%d/%m/%Y %H:%M')}</b>", subtitle_style))
            elements.append(Paragraph(f"Rango: <b>{context['start_date']} a {context['end_date']}</b>", subtitle_style))
            elements.append(Spacer(1, 10))
            # Tabla de ingresos
            table_data = [[
                Paragraph('<b>Fecha</b>', styles['Normal']),
                Paragraph('<b>Placa</b>', styles['Normal']),
                Paragraph('<b>Categoría</b>', styles['Normal']),
                Paragraph('<b>Monto</b>', styles['Normal']),
                Paragraph('<b>Forma de Pago</b>', styles['Normal']),
            ]]
            total = 0
            pagos = {f['clave']: 0 for f in context['formas_pago_list']}
            for ticket in ParkingTicket.objects.filter(
                tenant=request.tenant,
                exit_time__isnull=False,
                exit_time__range=(datetime.combine(context['start_date'], datetime.min.time()), datetime.combine(context['end_date'], datetime.max.time()))
            ).exclude(amount_paid__isnull=True):
                forma_pago = ticket.forma_pago or 'efectivo'
                if forma_pago not in pagos:
                    forma_pago = 'efectivo'
                pagos[forma_pago] += ticket.amount_paid
                total += ticket.amount_paid
                table_data.append([
                    ticket.exit_time.strftime('%d/%m/%Y %H:%M'),
                    ticket.placa,
                    ticket.category.name if ticket.category else '',
                    Paragraph(f"<para alignment='right'><b>${ticket.amount_paid:,.2f}</b></para>", styles['Normal']),
                    Paragraph(f"<para alignment='center'>{dict(CajaMovimiento.FORMA_PAGO_CHOICES).get(forma_pago, forma_pago.title())}</para>", styles['Normal']),
                ])
            col_widths = [90, 80, 100, 90, 120]
            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (2,-1), 'CENTER'),
                ('ALIGN', (3,1), (3,-1), 'RIGHT'),
                ('ALIGN', (4,1), (4,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 11),
                ('FONTSIZE', (0,1), (-1,-1), 10),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.HexColor('#f1f5f9')]),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('BOTTOMPADDING', (0,0), (-1,0), 8),
                ('TOPPADDING', (0,0), (-1,0), 8),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 16))
            # Layout profesional: desglose y total alineados horizontalmente
            total_style = ParagraphStyle('total', parent=styles['Heading2'], fontSize=16, textColor=colors.HexColor('#16a34a'), alignment=TA_RIGHT, spaceAfter=0)
            desglose_style = ParagraphStyle('desglose', parent=styles['Heading3'], fontSize=13, textColor=colors.HexColor('#2563eb'), alignment=TA_LEFT, spaceAfter=6)
            desglose_data = [[Paragraph('<b>Medio de Pago</b>', styles['Normal']), Paragraph('<b>Monto</b>', styles['Normal'])]]
            for f in context['formas_pago_list']:
                desglose_data.append([
                    Paragraph(f["nombre"], styles['Normal']),
                    Paragraph(f"<para alignment='right'><b>${pagos[f['clave']]:,.2f}</b></para>", styles['Normal'])
                ])
            desglose_table = Table(desglose_data, colWidths=[140, 100])
            desglose_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 11),
                ('FONTSIZE', (0,1), (-1,-1), 10),
                ('ALIGN', (0,0), (0,-1), 'LEFT'),
                ('ALIGN', (1,1), (1,-1), 'RIGHT'),
                ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#cbd5e1')),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('TOPPADDING', (0,0), (-1,0), 6),
            ]))
            # Tabla de layout horizontal
            layout_table = Table([
                [Paragraph("Desglose por Medio de Pago:", desglose_style), Paragraph(f"TOTAL INGRESOS: <b>${total:,.2f}</b>", total_style)],
                [desglose_table, ""]
            ], colWidths=[260, 260])
            layout_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('ALIGN', (1,0), (1,0), 'RIGHT'),
                ('SPAN', (1,1), (1,1)),
                ('LEFTPADDING', (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING', (0,0), (-1,-1), 0),
                ('BOTTOMPADDING', (0,0), (-1,-1), 0),
            ]))
            elements.append(layout_table)
            doc.build(elements)
            buffer.seek(0)
            return HttpResponse(buffer, content_type='application/pdf')
        return super().render_to_response(context, **response_kwargs)


@login_required
@tenant_required
@admin_required
def company_profile(request):
    tenant = request.tenant

    from .forms import TenantEditForm

    if request.method == 'POST':
        form = TenantEditForm(request.POST, instance=tenant)
        if form.is_valid():
            form.save()
            messages.success(request, 'Información de la empresa actualizada correctamente.')
            return redirect('company_profile')
    else:
        form = TenantEditForm(instance=tenant)

    return render(request, 'parking/company_profile.html', {'form': form})


def validate_plate(request, plate):
    exists = ParkingTicket.objects.filter(
        tenant=request.tenant,
        placa__iexact=plate,
        exit_time__isnull=True
    ).exists()
    return JsonResponse({'exists': exists})


@method_decorator(login_required, name='dispatch')
@method_decorator(tenant_required, name='dispatch')
@method_decorator(admin_required, name='dispatch')
class CategoryDeleteView(DeleteView):
    model = VehicleCategory
    success_url = reverse_lazy('category-list')
    template_name = 'parking/category_confirm_delete.html'

    def get_queryset(self):
        return VehicleCategory.objects.filter(tenant=self.request.tenant)

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Categoría eliminada exitosamente')
        return super().delete(request, *args, **kwargs)

@login_required
@tenant_required
@cajero_required
def cash_register(request):
    # Redirigir a la nueva vista profesional de caja
    from django.shortcuts import redirect
    return redirect('historial_cuadre_caja')

# Vistas para Superadministrador
@login_required
def superadmin_dashboard(request):
    """Dashboard del superadministrador"""
    if not request.user.is_superuser:
        messages.error(request, 'Acceso denegado. Solo superusuarios pueden acceder.')
        return redirect('dashboard')
    
    tenants = Tenant.objects.all().order_by('-created_at')
    active_tenants = tenants.filter(is_active=True).count()
    total_tenants = tenants.count()
    
    context = {
        'tenants': tenants,
        'active_tenants': active_tenants,
        'total_tenants': total_tenants,
    }
    return render(request, 'parking/superadmin/dashboard.html', context)


@login_required
def tenant_list(request):
    """Lista de todas las empresas"""
    if not request.user.is_superuser:
        messages.error(request, 'Acceso denegado. Solo superusuarios pueden acceder.')
        return redirect('dashboard')
    
    tenants = Tenant.objects.all().order_by('-created_at')
    
    context = {
        'tenants': tenants,
    }
    return render(request, 'parking/superadmin/tenant_list.html', context)


@login_required
def tenant_create(request):
    """Crear nueva empresa"""
    if not request.user.is_superuser:
        messages.error(request, 'Acceso denegado. Solo superusuarios pueden acceder.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = TenantCreateForm(request.POST)
        if form.is_valid():
            try:
                tenant = form.save()
                
                # Crear usuario administrador para la empresa
                admin_username = form.cleaned_data.get('admin_username')
                admin_email = form.cleaned_data.get('admin_email')
                admin_password = form.cleaned_data.get('admin_password')
                
                if admin_username and admin_email and admin_password:
                    try:
                        # Verificar si el usuario ya existe
                        if User.objects.filter(username=admin_username).exists():
                            messages.warning(request, f'El usuario "{admin_username}" ya existe. La empresa se creó pero no se pudo crear el usuario administrador.')
                        else:
                            # Crear usuario de Django
                            admin_user = User.objects.create_user(
                                username=admin_username,
                                email=admin_email,
                                password=admin_password
                            )
                            
                            # Crear perfil de administrador
                            profile = UserProfile.objects.create(
                                user=admin_user,
                                tenant=tenant,
                                role='admin',
                                is_active=True
                            )
                            
                            messages.success(request, f'Empresa "{tenant.name}" creada exitosamente con usuario administrador "{admin_username}".')
                    except Exception as e:
                        messages.warning(request, f'Empresa creada pero hubo un error al crear el usuario administrador: {str(e)}')
                else:
                    messages.success(request, f'Empresa "{tenant.name}" creada exitosamente.')
                
                return redirect('superadmin-tenant-list')
            except Exception as e:
                messages.error(request, f'Error al crear empresa: {str(e)}')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = TenantCreateForm()
    
    context = {
        'form': form,
        'title': 'Crear Nueva Empresa',
    }
    return render(request, 'parking/superadmin/tenant_form.html', context)


@login_required
def tenant_edit(request, pk):
    """Editar empresa existente"""
    if not request.user.is_superuser:
        messages.error(request, 'Acceso denegado. Solo superusuarios pueden acceder.')
        return redirect('dashboard')
    
    tenant = get_object_or_404(Tenant, pk=pk)
    
    if request.method == 'POST':
        form = TenantEditForm(request.POST, instance=tenant)
        if form.is_valid():
            tenant = form.save()
            messages.success(request, f'Empresa "{tenant.name}" actualizada exitosamente.')
            return redirect('superadmin-tenant-list')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = TenantEditForm(instance=tenant)
    
    context = {
        'form': form,
        'tenant': tenant,
        'title': f'Editar Empresa: {tenant.name}',
    }
    return render(request, 'parking/superadmin/tenant_form.html', context)


@login_required
def tenant_detail(request, pk):
    """Detalles de una empresa"""
    if not request.user.is_superuser:
        messages.error(request, 'Acceso denegado. Solo superusuarios pueden acceder.')
        return redirect('dashboard')
    
    tenant = get_object_or_404(Tenant, pk=pk)
    
    # Estadísticas de la empresa
    total_tickets = ParkingTicket.objects.filter(tenant=tenant).count()
    active_tickets = ParkingTicket.objects.filter(tenant=tenant, exit_time__isnull=True).count()
    total_revenue = ParkingTicket.objects.filter(
        tenant=tenant, 
        exit_time__isnull=False,
        amount_paid__isnull=False
    ).aggregate(total=Sum('amount_paid'))['total'] or 0
    
    context = {
        'tenant': tenant,
        'total_tickets': total_tickets,
        'active_tickets': active_tickets,
        'total_revenue': total_revenue,
    }
    return render(request, 'parking/superadmin/tenant_detail.html', context)


@login_required
def tenant_toggle_status(request, pk):
    """Activar/desactivar empresa"""
    if not request.user.is_superuser:
        messages.error(request, 'Acceso denegado. Solo superusuarios pueden acceder.')
        return redirect('dashboard')
    
    tenant = get_object_or_404(Tenant, pk=pk)
    
    if request.method == 'POST':
        tenant.is_active = not tenant.is_active
        tenant.save()
        
        status = "activada" if tenant.is_active else "desactivada"
        messages.success(request, f'Empresa "{tenant.name}" {status} exitosamente.')
        
        return redirect('superadmin-tenant-list')
    
    return redirect('superadmin-tenant-detail', pk=pk)


@login_required
def tenant_login(request, pk):
    """Login como empresa específica (para superadmin)"""
    if not request.user.is_superuser:
        messages.error(request, 'Acceso denegado. Solo superusuarios pueden acceder.')
        return redirect('dashboard')
    
    tenant = get_object_or_404(Tenant, pk=pk)
    
    # Almacenar el tenant en la sesión para acceso temporal
    request.session['temp_tenant_id'] = tenant.id
    request.session['temp_tenant_name'] = tenant.name
    
    messages.success(request, f'Has iniciado sesión como administrador de "{tenant.name}"')
    return redirect('dashboard')

@login_required
def tenant_selection(request):
    """Vista para seleccionar empresa (login multitenant)"""
    tenants = Tenant.objects.filter(is_active=True).order_by('name')
    
    context = {
        'tenants': tenants,
    }
    return render(request, 'parking/superadmin/tenant_login.html', context)


@login_required
def exit_tenant_mode(request):
    """Salir del modo empresa y volver al panel de superadmin"""
    if not request.user.is_superuser:
        messages.error(request, 'Acceso denegado. Solo superusuarios pueden acceder.')
        return redirect('dashboard')
    
    # Limpiar tenant temporal de la sesión
    request.session.pop('temp_tenant_id', None)
    request.session.pop('temp_tenant_name', None)
    
    messages.success(request, 'Has salido del modo empresa y vuelto al panel de administración.')
    return redirect('superadmin-dashboard')

# Vistas para Gestión de Usuarios de Empresa
@login_required
@tenant_required
@admin_required
def tenant_users_list(request):
    """Lista de usuarios de la empresa"""
    users = UserProfile.objects.filter(tenant=request.tenant).select_related('user')
    
    context = {
        'users': users,
    }
    return render(request, 'parking/tenant/users_list.html', context)


@login_required
@tenant_required
@admin_required
def tenant_user_create(request):
    """Crear nuevo usuario para la empresa"""
    tenant = request.tenant
    if request.method == 'POST':
        form = TenantUserForm(request.POST, tenant=tenant)
        if form.is_valid():
            try:
                # Usar el método save del formulario que maneja todo
                profile = form.save()
                messages.success(request, f'Usuario "{profile.user.username}" creado exitosamente.')
                return redirect('tenant-users-list')
            except Exception as e:
                messages.error(request, f'Error al crear usuario: {str(e)}')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = TenantUserForm(tenant=tenant)
    
    context = {
        'form': form,
        'title': 'Crear Nuevo Usuario',
    }
    return render(request, 'parking/tenant/user_form.html', context)


@login_required
@tenant_required
@admin_required
def tenant_user_edit(request, pk):
    """Editar usuario de la empresa"""
    profile = get_object_or_404(UserProfile, pk=pk, tenant=request.tenant)
    
    if request.method == 'POST':
        form = TenantUserEditForm(request.POST, instance=profile, tenant=request.tenant)
        if form.is_valid():
            form.save()
            messages.success(request, f'Usuario "{profile.user.username}" actualizado exitosamente.')
            return redirect('tenant-users-list')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = TenantUserEditForm(instance=profile, tenant=request.tenant)
    
    context = {
        'form': form,
        'user_profile': profile,
        'title': f'Editar Usuario: {profile.user.username}',
    }
    return render(request, 'parking/tenant/user_form.html', context)


@login_required
@tenant_required
@admin_required
def tenant_user_password(request, pk):
    """Cambiar contraseña de usuario"""
    profile = get_object_or_404(UserProfile, pk=pk, tenant=request.tenant)
    
    if request.method == 'POST':
        form = TenantUserPasswordForm(request.POST)
        if form.is_valid():
            profile.user.set_password(form.cleaned_data['password1'])
            profile.user.save()
            messages.success(request, f'Contraseña de "{profile.user.username}" cambiada exitosamente.')
            return redirect('tenant-users-list')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = TenantUserPasswordForm()
    
    context = {
        'form': form,
        'user_profile': profile,
        'title': f'Cambiar Contraseña: {profile.user.username}',
    }
    return render(request, 'parking/tenant/user_password.html', context)


@login_required
@tenant_required
@admin_required
def tenant_user_toggle_status(request, pk):
    """Activar/desactivar usuario de la empresa"""
    profile = get_object_or_404(UserProfile, pk=pk, tenant=request.tenant)
    
    if request.method == 'POST':
        profile.is_active = not profile.is_active
        profile.save()
        
        status = "activado" if profile.is_active else "desactivado"
        messages.success(request, f'Usuario "{profile.user.username}" {status} exitosamente.')
        
        return redirect('tenant-users-list')
    
    return redirect('tenant-users-list')


@login_required
@tenant_required
@admin_required
def tenant_user_delete(request, pk):
    """Eliminar usuario de la empresa"""
    profile = get_object_or_404(UserProfile, pk=pk, tenant=request.tenant)
    
    if request.method == 'POST':
        username = profile.user.username
        profile.user.delete()  # Esto también elimina el perfil por CASCADE
        messages.success(request, f'Usuario "{username}" eliminado exitosamente.')
        return redirect('tenant-users-list')
    
    context = {
        'user_profile': profile,
    }
    return render(request, 'parking/tenant/user_confirm_delete.html', context)

@login_required
def tenant_inactive(request):
    """Vista para mostrar cuando una empresa está inactiva"""
    tenant = getattr(request, 'tenant', None)
    return render(request, 'parking/tenant_inactive.html', {
        'tenant': tenant,
        'tenant_name': tenant.name if tenant else 'Empresa'
    })

@login_required
@tenant_required
@cajero_required
def apertura_turno(request):
    if CajaTurno.objects.filter(tenant=request.tenant, estado='abierto').exists():
        messages.warning(request, 'Ya hay un turno de caja abierto. Debe cerrarlo antes de abrir uno nuevo.')
        return redirect('historial_cuadre_caja')
    if request.method == 'POST':
        form = CajaTurnoAperturaForm(request.POST)
        if form.is_valid():
            turno = form.save(commit=False)
            turno.tenant = request.tenant
            turno.usuario_apertura = request.user
            turno.estado = 'abierto'
            turno.save()
            messages.success(request, 'Turno de caja abierto correctamente.')
            return redirect('historial_cuadre_caja')
    else:
        form = CajaTurnoAperturaForm()
    return render(request, 'parking/caja/apertura_turno.html', {'form': form})

@login_required
@tenant_required
@cajero_required
def registrar_movimiento(request):
    turno = CajaTurno.objects.filter(tenant=request.tenant, estado='abierto').first()
    if not turno:
        messages.error(request, 'No hay un turno de caja abierto. Debe abrir uno antes de registrar movimientos.')
        return redirect('apertura_turno')
    if request.method == 'POST':
        form = CajaMovimientoForm(request.POST)
        if form.is_valid():
            movimiento = form.save(commit=False)
            movimiento.caja_turno = turno
            movimiento.usuario = request.user
            movimiento.save()
            messages.success(request, 'Movimiento registrado correctamente.')
            return redirect('historial_cuadre_caja')
    else:
        form = CajaMovimientoForm()
    return render(request, 'parking/caja/registrar_movimiento.html', {'form': form, 'turno': turno})

@login_required
@tenant_required
@cajero_required
def cierre_turno(request):
    turno = CajaTurno.objects.filter(tenant=request.tenant, estado='abierto').first()
    if not turno:
        messages.error(request, 'No hay un turno de caja abierto para cerrar.')
        return redirect('historial_cuadre_caja')
    if request.method == 'POST':
        form = CajaTurnoCierreForm(request.POST, instance=turno)
        if form.is_valid():
            cierre = form.save(commit=False)
            cierre.usuario_cierre = request.user
            cierre.fecha_cierre = timezone.now()
            cierre.estado = 'cerrado'
            cierre.save()
            messages.success(request, 'Turno de caja cerrado correctamente.')
            return redirect('historial_cuadre_caja')
    else:
        form = CajaTurnoCierreForm(instance=turno)
    resumen = {
        'dinero_inicial_en_caja': turno.monto_inicial,  # Dinero inicial en caja
        'total_ingresos': turno.total_ingresos(),
        'total_egresos': turno.total_egresos(),
        'total_esperado': turno.total_esperado(),
        'total_esperado_efectivo': turno.total_esperado_efectivo(),
        'diferencia': turno.diferencia_efectivo(),
        'alerta': turno.alerta_diferencia(),
        'desglose_forma_pago': turno.desglose_ingresos_por_forma_pago(),
    }
    return render(request, 'parking/caja/cierre_turno.html', {'form': form, 'turno': turno, 'resumen': resumen})

@login_required
@tenant_required
@cajero_required
def historial_cuadre_caja(request):
    turnos = CajaTurno.objects.filter(tenant=request.tenant).order_by('-fecha_apertura')
    # Filtros por fecha, usuario, diferencia, forma de pago
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    usuario = request.GET.get('usuario')
    diferencia = request.GET.get('diferencia')
    forma_pago = request.GET.get('forma_pago')
    if fecha_inicio:
        turnos = turnos.filter(fecha_apertura__date__gte=fecha_inicio)
    if fecha_fin:
        turnos = turnos.filter(fecha_cierre__date__lte=fecha_fin)
    if usuario:
        turnos = turnos.filter(usuario_apertura__username__icontains=usuario)
    if diferencia == 'si':
        turnos = [t for t in turnos if t.diferencia() and abs(t.diferencia()) > 0]
    elif diferencia == 'no':
        turnos = [t for t in turnos if not t.diferencia() or abs(t.diferencia()) == 0]
    if forma_pago:
        turnos = [t for t in turnos if t.movimientos.filter(forma_pago=forma_pago).exists()]
    # Exportar a PDF
    if 'exportar_pdf' in request.GET:
        try:
            # Verificar permisos del usuario
            if not request.user.is_authenticated:
                return JsonResponse({'error': 'Debes iniciar sesión para exportar el PDF.'}, status=401)
            
            # Verificar que el usuario tenga rol de cajero o admin
            try:
                user_profile = request.user.profile
                if user_profile.role not in ['admin', 'cajero']:
                    return JsonResponse({'error': 'Solo usuarios con rol de cajero o administrador pueden exportar PDF.'}, status=403)
            except:
                return JsonResponse({'error': 'Perfil de usuario no encontrado.'}, status=400)
            
            # Verificar que haya tenant asignado
            if not hasattr(request, 'tenant') or not request.tenant:
                return JsonResponse({'error': 'No tienes una empresa asignada.'}, status=400)
            
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
            elements = []
            styles = getSampleStyleSheet()
            title = Paragraph("<b>Historial de Cuadre de Caja</b>", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 8))
            info = Paragraph(f"<b>Fecha de generación:</b> {timezone.now().strftime('%d/%m/%Y %H:%M')}<br/><b>Empresa:</b> {request.tenant.name}<br/><b>Usuario:</b> {request.user.username}", styles['Normal'])
            elements.append(info)
            elements.append(Spacer(1, 8))

            # Tabla principal
            headers = [
                "Apertura", "Cierre", "Usr. Ap.", "Usr. Ci.", "Inicial", "Efec.", "Dat.", "Trans.", "Egr.", "Esp. Efec.", "Contado", "Dif.", "Estado"
            ]
            data = [headers]
            for turno in turnos:
                ingresos_por_forma = turno.desglose_ingresos_por_forma_pago()
                data.append([
                    turno.fecha_apertura.strftime('%d/%m/%y %H:%M'),
                    turno.fecha_cierre.strftime('%d/%m/%y %H:%M') if turno.fecha_cierre else '-',
                    getattr(turno.usuario_apertura, 'username', '-')[:10],
                    getattr(turno.usuario_cierre, 'username', '-')[:10] if turno.usuario_cierre else '-',
                    f"${turno.monto_inicial:,.2f}",
                    f"${ingresos_por_forma.get('efectivo', 0):,.2f}",
                    f"${ingresos_por_forma.get('datafono', 0):,.2f}",
                    f"${ingresos_por_forma.get('transferencia', 0):,.2f}",
                    f"${turno.total_egresos():,.2f}",
                    f"${turno.total_esperado_efectivo():,.2f}",
                    f"${turno.monto_real_contado:,.2f}" if turno.monto_real_contado else '-',
                    f"${turno.diferencia_efectivo():,.2f}" if turno.diferencia_efectivo() is not None else '-',
                    turno.get_estado_display(),
                ])

            col_widths = [60, 60, 55, 55, 50, 50, 50, 50, 50, 60, 60, 50, 50]
            t = Table(data, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2563eb')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 8),
                ('FONTSIZE', (0,1), (-1,-1), 7),
                ('BOTTOMPADDING', (0,0), (-1,0), 6),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 12))

            # Resumen final discriminado por tipo de pago
            if turnos:
                total_inicial = sum(t.monto_inicial for t in turnos)
                total_efectivo = sum(t.desglose_ingresos_por_forma_pago().get('efectivo', 0) for t in turnos)
                total_datafono = sum(t.desglose_ingresos_por_forma_pago().get('datafono', 0) for t in turnos)
                total_transferencia = sum(t.desglose_ingresos_por_forma_pago().get('transferencia', 0) for t in turnos)
                total_egresos = sum(t.total_egresos() for t in turnos)
                resumen_data = [
                    ["Total Inicial", f"${total_inicial:,.2f}"],
                    ["Total Efectivo", f"${total_efectivo:,.2f}"],
                    ["Total Datafono", f"${total_datafono:,.2f}"],
                    ["Total Transferencia", f"${total_transferencia:,.2f}"],
                    ["Total Egresos", f"${total_egresos:,.2f}"],
                    ["Balance Neto", f"${(total_inicial + total_efectivo + total_datafono + total_transferencia) - total_egresos:,.2f}"],
                ]
                resumen_table = Table(resumen_data, colWidths=[120, 100])
                resumen_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
                    ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
                    ('FONTSIZE', (0,0), (-1,-1), 9),
                    ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                    ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
                ]))
                elements.append(resumen_table)

            doc.build(elements)
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="historial_cuadre_caja_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            return response

                # Detalle de movimientos
            
            # Resumen final discriminado por tipo de pago
            if turnos:
                y -= 20
                p.setStrokeColor(colors.HexColor('#2563eb'))
                p.setLineWidth(1)
                p.line(40, y, width-40, y)
                y -= 20

                total_inicial = sum(t.monto_inicial for t in turnos)
                total_efectivo = sum(t.desglose_ingresos_por_forma_pago().get('efectivo', 0) for t in turnos)
                total_datafono = sum(t.desglose_ingresos_por_forma_pago().get('datafono', 0) for t in turnos)
                total_transferencia = sum(t.desglose_ingresos_por_forma_pago().get('transferencia', 0) for t in turnos)
                total_otro = sum(t.desglose_ingresos_por_forma_pago().get('otro', 0) for t in turnos)
                total_egresos = sum(t.total_egresos() for t in turnos)

                p.setFont("Helvetica-Bold", 12)
                p.setFillColor(colors.HexColor('#2563eb'))
                p.drawString(40, y, f"TOTAL DINERO INICIAL: ${total_inicial:,.2f}")
                y -= 16
                p.setFillColor(colors.HexColor('#16a34a'))
                p.drawString(40, y, f"TOTAL INGRESOS EFECTIVO: ${total_efectivo:,.2f}")
                y -= 16
                p.setFillColor(colors.HexColor('#0ea5e9'))
                p.drawString(40, y, f"TOTAL INGRESOS DATAFONO: ${total_datafono:,.2f}")
                y -= 16
                p.setFillColor(colors.HexColor('#a21caf'))
                p.drawString(40, y, f"TOTAL INGRESOS TRANSFERENCIA: ${total_transferencia:,.2f}")
                y -= 16
                p.setFillColor(colors.HexColor('#f59e42'))
                p.drawString(40, y, f"TOTAL INGRESOS OTRO: ${total_otro:,.2f}")
                y -= 16
                p.setFillColor(colors.HexColor('#dc2626'))
                p.drawString(40, y, f"TOTAL EGRESOS: ${total_egresos:,.2f}")
                y -= 16
                p.setFillColor(colors.HexColor('#2563eb'))
                p.drawString(40, y, f"BALANCE NETO: ${(total_inicial + total_efectivo + total_datafono + total_transferencia + total_otro) - total_egresos:,.2f}")
            
            p.save()
            buffer.seek(0)
            
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="historial_cuadre_caja_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            return response
            
        except Exception as e:
            return JsonResponse({'error': f'Error al generar PDF: {str(e)}'}, status=500)
    # Calcular y asignar esperado_efectivo y diferencia como en el PDF
    for t in turnos:
        t.esperado_efectivo = t.total_esperado_efectivo()  # Solo efectivo: inicial + ingresos efectivo - egresos efectivo
        if t.monto_real_contado is not None:
            t.diferencia = t.monto_real_contado - t.total_esperado_efectivo()
        else:
            t.diferencia = None
    return render(request, 'parking/caja/historial_cuadre_caja.html', {'turnos': turnos})

# --- VISTAS PARA GESTIÓN DE ROLES Y PERMISOS ---

@login_required
@tenant_required
@permission_required('manage_roles')
def roles_list(request):
    """Lista de roles personalizados"""
    search_form = RoleSearchForm(request.GET)
    roles = CustomRole.objects.filter(tenant=request.tenant)
    
    if search_form.is_valid():
        search = search_form.cleaned_data.get('search')
        is_active = search_form.cleaned_data.get('is_active')
        
        if search:
            roles = roles.filter(name__icontains=search)
        
        if is_active:
            roles = roles.filter(is_active=(is_active == 'true'))
    
    roles = roles.order_by('name')
    
    context = {
        'roles': roles,
        'search_form': search_form,
    }
    return render(request, 'parking/roles/roles_list.html', context)


@login_required
@tenant_required
@permission_required('manage_roles')
def role_create(request):
    """Crear nuevo rol personalizado"""
    if request.method == 'POST':
        form = CustomRoleForm(request.POST, tenant=request.tenant)
        if form.is_valid():
            role = form.save()
            messages.success(request, f'Rol "{role.name}" creado exitosamente.')
            return redirect('role-permissions', pk=role.pk)
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = CustomRoleForm(tenant=request.tenant)
    
    context = {
        'form': form,
        'title': 'Crear Nuevo Rol',
    }
    return render(request, 'parking/roles/role_form.html', context)


@login_required
@tenant_required
@permission_required('manage_roles')
def role_edit(request, pk):
    """Editar rol personalizado"""
    role = get_object_or_404(CustomRole, pk=pk, tenant=request.tenant)
    
    if request.method == 'POST':
        form = CustomRoleForm(request.POST, instance=role, tenant=request.tenant)
        if form.is_valid():
            form.save()
            messages.success(request, f'Rol "{role.name}" actualizado exitosamente.')
            return redirect('roles-list')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = CustomRoleForm(instance=role, tenant=request.tenant)
    
    context = {
        'form': form,
        'role': role,
        'title': f'Editar Rol: {role.name}',
    }
    return render(request, 'parking/roles/role_form.html', context)


@login_required
@tenant_required
@permission_required('manage_roles')
def role_permissions(request, pk):
    """Asignar permisos a un rol"""
    role = get_object_or_404(CustomRole, pk=pk, tenant=request.tenant)
    
    if request.method == 'POST':
        form = RolePermissionForm(request.POST, role=role)
        if form.is_valid():
            # Eliminar permisos actuales del rol
            RolePermission.objects.filter(role=role).delete()
            
            # Asignar nuevos permisos
            for field_name, value in form.cleaned_data.items():
                if field_name.startswith('perm_') and value:
                    permission_code = field_name.replace('perm_', '')
                    try:
                        permission = Permission.objects.get(code=permission_code, is_active=True)
                        RolePermission.objects.create(role=role, permission=permission)
                    except Permission.DoesNotExist:
                        pass
            
            messages.success(request, f'Permisos del rol "{role.name}" actualizados exitosamente.')
            return redirect('roles-list')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = RolePermissionForm(role=role)
    
    context = {
        'form': form,
        'role': role,
        'title': f'Permisos del Rol: {role.name}',
    }
    return render(request, 'parking/roles/role_permissions.html', context)


@login_required
@tenant_required
@permission_required('manage_roles')
def role_delete(request, pk):
    """Eliminar rol personalizado"""
    role = get_object_or_404(CustomRole, pk=pk, tenant=request.tenant)
    
    if request.method == 'POST':
        role_name = role.name
        role.delete()
        messages.success(request, f'Rol "{role_name}" eliminado exitosamente.')
        return redirect('roles-list')
    
    context = {
        'role': role,
    }
    return render(request, 'parking/roles/role_confirm_delete.html', context)


@login_required
@tenant_required
@permission_required('manage_roles')
def role_toggle_status(request, pk):
    """Activar/desactivar rol"""
    role = get_object_or_404(CustomRole, pk=pk, tenant=request.tenant)
    
    if request.method == 'POST':
        role.is_active = not role.is_active
        role.save()
        
        status = "activado" if role.is_active else "desactivado"
        messages.success(request, f'Rol "{role.name}" {status} exitosamente.')
        
        return redirect('roles-list')
    
    return redirect('roles-list')


@login_required
@tenant_required
@permission_required('edit_users')
def user_roles(request, pk):
    """Asignar roles a un usuario"""
    user_profile = get_object_or_404(UserProfile, pk=pk, tenant=request.tenant)
    
    if request.method == 'POST':
        form = UserRoleForm(request.POST, user_profile=user_profile)
        if form.is_valid():
            # Eliminar roles actuales del usuario
            UserRole.objects.filter(user_profile=user_profile).delete()
            
            # Asignar nuevos roles
            primary_role_id = form.cleaned_data.get('primary_role')
            
            for field_name, value in form.cleaned_data.items():
                if field_name.startswith('role_') and value:
                    role_id = field_name.replace('role_', '')
                    try:
                        role = CustomRole.objects.get(id=role_id, tenant=request.tenant, is_active=True)
                        is_primary = str(role.id) == primary_role_id
                        UserRole.objects.create(
                            user_profile=user_profile,
                            role=role,
                            is_primary=is_primary
                        )
                    except CustomRole.DoesNotExist:
                        pass
            
            messages.success(request, f'Roles del usuario "{user_profile.user.username}" actualizados exitosamente.')
            return redirect('tenant-users-list')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = UserRoleForm(user_profile=user_profile)
    
    context = {
        'form': form,
        'user_profile': user_profile,
        'title': f'Roles del Usuario: {user_profile.user.username}',
    }
    return render(request, 'parking/roles/user_roles.html', context)


@login_required
@tenant_required
@permission_required('edit_users')
def user_permissions(request, pk):
    """Asignar permisos específicos a un usuario"""
    user_profile = get_object_or_404(UserProfile, pk=pk, tenant=request.tenant)
    
    if request.method == 'POST':
        form = UserPermissionForm(request.POST, user_profile=user_profile)
        if form.is_valid():
            # Eliminar permisos específicos actuales del usuario
            UserPermission.objects.filter(user_profile=user_profile).delete()
            
            # Asignar nuevos permisos específicos
            for field_name, value in form.cleaned_data.items():
                if field_name.startswith('perm_'):
                    permission_code = field_name.replace('perm_', '')
                    try:
                        permission = Permission.objects.get(code=permission_code, is_active=True)
                        # Solo crear registro si es diferente del permiso por rol
                        has_by_role = user_profile.has_permission(permission_code)
                        if value != has_by_role:
                            UserPermission.objects.create(
                                user_profile=user_profile,
                                permission=permission,
                                is_granted=value
                            )
                    except Permission.DoesNotExist:
                        pass
            
            messages.success(request, f'Permisos del usuario "{user_profile.user.username}" actualizados exitosamente.')
            return redirect('tenant-users-list')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = UserPermissionForm(user_profile=user_profile)
    
    context = {
        'form': form,
        'user_profile': user_profile,
        'title': f'Permisos del Usuario: {user_profile.user.username}',
    }
    return render(request, 'parking/roles/user_permissions.html', context)


@login_required
@tenant_required
@permission_required('view_users')
def permissions_overview(request):
    """Vista general de permisos y usuarios"""
    # Obtener todos los usuarios del tenant
    users = UserProfile.objects.filter(tenant=request.tenant).select_related('user')
    
    # Obtener todos los roles del tenant
    roles = CustomRole.objects.filter(tenant=request.tenant, is_active=True)
    
    # Obtener todos los módulos
    modules = Module.objects.filter(is_active=True).order_by('order')
    
    # Crear matriz de permisos
    permissions_matrix = []
    for user in users:
        user_permissions = user.get_permissions()
        user_roles = user.get_roles()
        
        permissions_matrix.append({
            'user': user,
            'roles': user_roles,
            'permissions': user_permissions,
            'modules': user.get_modules()
        })
    
    context = {
        'users': users,
        'roles': roles,
        'modules': modules,
        'permissions_matrix': permissions_matrix,
    }
    return render(request, 'parking/roles/permissions_overview.html', context)


@login_required
@tenant_required
@permission_required('view_users')
def access_log(request):
    """Bitácora de acceso y cambios"""
    logs = AccessLog.objects.filter(tenant=request.tenant).select_related('user').order_by('-created_at')
    
    # Filtros
    action = request.GET.get('action')
    user_id = request.GET.get('user')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if action:
        logs = logs.filter(action=action)
    
    if user_id:
        logs = logs.filter(user_id=user_id)
    
    if date_from:
        logs = logs.filter(created_at__date__gte=date_from)
    
    if date_to:
        logs = logs.filter(created_at__date__lte=date_to)
    
    # Paginación
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'action_choices': AccessLog.ACTION_CHOICES,
        'users': UserProfile.objects.filter(tenant=request.tenant).select_related('user'),
    }
    return render(request, 'parking/roles/access_log.html', context)